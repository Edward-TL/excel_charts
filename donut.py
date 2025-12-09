```python
"""donut.py

Implementation of a donut (pie) chart using OpenPyXL.
"""

from __future__ import annotations

from openpyxl.chart import PieChart, Reference

from .core import BaseChart, MoneyAxis


class DonutChartWrapper(BaseChart):
    """Donut chart wrapper.

    OpenPyXL does not have a dedicated ``DonutChart`` class, but a ``PieChart``
    can be styled to appear as a donut by setting the ``holeSize`` attribute.
    This implementation writes the DataFrame to the worksheet and creates a
    ``PieChart`` with a hole.
    """

    def _create_chart(self):
        chart = PieChart()
        chart.title = "Donut Chart"
        chart.style = 10
        # ``holeSize`` is a percentage (0‑100). 50 gives a classic donut look.
        if hasattr(chart, "holeSize"):
            chart.holeSize = 50
        else:
            # Fallback – some older versions of openpyxl ignore ``holeSize``.
            pass

        # For a donut we treat the first column as categories and the second as values.
        data = self.data
        # Write data to worksheet
        for r_idx, row in enumerate(data.itertuples(index=False), start=2):
            self._ws.cell(row=r_idx, column=1, value=getattr(row, data.columns[0]))
            self._ws.cell(row=r_idx, column=2, value=getattr(row, data.columns[1]))
        # Header row
        self._ws.cell(row=1, column=1, value=data.columns[0])
        self._ws.cell(row=1, column=2, value=data.columns[1])

        min_row = 1
        max_row = len(data) + 1
        labels = Reference(self._ws, min_col=1, min_row=min_row, max_row=max_row)
        values = Reference(self._ws, min_col=2, min_row=min_row, max_row=max_row)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(labels)
        return chart
```
