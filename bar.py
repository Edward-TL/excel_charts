```python
"""bar.py

Implementation of vertical and horizontal bar charts using OpenPyXL.
"""

from __future__ import annotations

from enum import Enum

from openpyxl.chart import BarChart, Reference

from .core import BaseChart, MoneyAxis


class BarOrientation(str, Enum):
    """Orientation of the bar chart."""

    VERTICAL = "vertical"
    HORIZONTAL = "horizontal"


class BarChartWrapper(BaseChart):
    """Bar chart wrapper.

    Parameters
    ----------
    orientation : BarOrientation, optional
        Determines whether the bars are drawn vertically (default) or horizontally.
    """

    def __init__(self, *args, orientation: BarOrientation = BarOrientation.VERTICAL, **kwargs):
        super().__init__(*args, **kwargs)
        self.orientation = orientation

    def _create_chart(self):
        chart = BarChart()
        chart.title = "Bar Chart"
        chart.style = 11
        chart.y_axis.title = "Y Axis"
        chart.x_axis.title = "X Axis"

        # Set orientation – OpenPyXL uses ``type`` attribute for bar direction.
        if self.orientation == BarOrientation.HORIZONTAL:
            chart.type = "bar"
            chart.bar_dir = "col"
        else:
            chart.type = "col"
            chart.bar_dir = "col"

        # Determine which column is the money axis
        if self.money_axis == MoneyAxis.Y:
            y_col = 2
            x_col = 1
        else:
            y_col = 1
            x_col = 2

        data = self.data
        # Write DataFrame to worksheet
        for r_idx, row in enumerate(data.itertuples(index=False), start=2):
            self._ws.cell(row=r_idx, column=1, value=getattr(row, data.columns[x_col - 1]))
            self._ws.cell(row=r_idx, column=2, value=getattr(row, data.columns[y_col - 1]))

        # Header row
        self._ws.cell(row=1, column=1, value=data.columns[x_col - 1])
        self._ws.cell(row=1, column=2, value=data.columns[y_col - 1])

        min_row = 1
        max_row = len(data) + 1
        cats = Reference(self._ws, min_col=1, min_row=min_row, max_row=max_row)
        vals = Reference(self._ws, min_col=2, min_row=min_row, max_row=max_row)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        return chart
```
