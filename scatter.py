```python
"""scatter.py

Implementation of a scatter chart using OpenPyXL.
"""

from __future__ import annotations

from openpyxl.chart import ScatterChart, Reference, Series

from .core import BaseChart, MoneyAxis


class ScatterChartWrapper(BaseChart):
    """Scatter chart wrapper.

    The DataFrame should contain two numeric columns. The ``money_axis``
    determines which column is considered the monetary value.
    """

    def _create_chart(self):
        chart = ScatterChart()
        chart.title = "Scatter Chart"
        chart.style = 13
        chart.x_axis.title = "X Axis"
        chart.y_axis.title = "Y Axis"

        # Determine column mapping based on money_axis
        if self.money_axis == MoneyAxis.Y:
            x_col = 1
            y_col = 2
        else:
            x_col = 2
            y_col = 1

        data = self.data
        # Write DataFrame values to worksheet
        for r_idx, row in enumerate(data.itertuples(index=False), start=2):
            self._ws.cell(row=r_idx, column=1, value=getattr(row, data.columns[x_col - 1]))
            self._ws.cell(row=r_idx, column=2, value=getattr(row, data.columns[y_col - 1]))

        # Header row
        self._ws.cell(row=1, column=1, value=data.columns[x_col - 1])
        self._ws.cell(row=1, column=2, value=data.columns[y_col - 1])

        min_row = 1
        max_row = len(data) + 1
        x_ref = Reference(self._ws, min_col=1, min_row=min_row, max_row=max_row)
        y_ref = Reference(self._ws, min_col=2, min_row=min_row, max_row=max_row)
        series = Series(y_ref, x_ref, title_from_data=True)
        chart.series.append(series)
        return chart
```
