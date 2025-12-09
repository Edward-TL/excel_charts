```python
"""line.py

Implementation of a line chart using OpenPyXL.
"""

from __future__ import annotations

from openpyxl.chart import LineChart, Reference

from .core import BaseChart, MoneyAxis


class LineChartWrapper(BaseChart):
    """Line chart wrapper.

    Expects the DataFrame to have at least two columns: the first column for the
    X‑axis values and the second column for the Y‑axis values.
    """

    def _create_chart(self):
        chart = LineChart()
        chart.title = "Line Chart"
        chart.style = 12
        chart.y_axis.title = "Y Axis"
        chart.x_axis.title = "X Axis"

        # Determine which column is the money axis
        if self.money_axis == MoneyAxis.Y:
            y_col = 2
            x_col = 1
        else:
            y_col = 1
            x_col = 2

        data = self.data
        # OpenPyXL expects a worksheet with data; we write the DataFrame to the sheet first.
        for r_idx, row in enumerate(data.itertuples(index=False), start=2):
            # Write X value
            self._ws.cell(row=r_idx, column=1, value=getattr(row, data.columns[x_col - 1]))
            # Write Y value
            self._ws.cell(row=r_idx, column=2, value=getattr(row, data.columns[y_col - 1]))

        # Header row
        self._ws.cell(row=1, column=1, value=data.columns[x_col - 1])
        self._ws.cell(row=1, column=2, value=data.columns[y_col - 1])

        # Define data range for the chart
        min_row = 1
        max_row = len(data) + 1
        x_ref = Reference(self._ws, min_col=1, min_row=min_row, max_row=max_row)
        y_ref = Reference(self._ws, min_col=2, min_row=min_row, max_row=max_row)
        chart.add_data(y_ref, titles_from_data=True)
        chart.set_categories(x_ref)
        return chart
```
